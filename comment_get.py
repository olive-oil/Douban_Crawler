import os
import random
import re
import time
import traceback
from lxml import etree
import openpyxl as openpyxl
import requests
from bs4 import BeautifulSoup
from selenium.webdriver.chrome import webdriver


class DB_Film_review_Spider(object):
    def __init__(self):
        self.page=0
        self.total_author=[]
        self.total_star_num=[]
        self.total_comment_time=[]
        self.total_recommendation_level=[]
        self.total_text=[]
        self.level=['力荐','推荐','还行','较差','很差']
        self.num=[0,0,0,0,0]
        self.COMMENT_FILE_PATH = '深夜食堂豆瓣影评.xlsx'

    def loginDou(self,url,cookies,headers):
        """
        用cookie登录豆瓣
        :param url:
        :param cookies:
        :param headers:
        :return:
        """
        session.headers = headers
        requests.utils.add_dict_to_cookiejar(session.cookies, cookies)
        try:
            r = session.get(url)
            r.raise_for_status()
            print('login douban successfully')
            return 1
        except:
            print("failure with session login")
            return 0

    def getComment(self,commentUrl,comHeaders,start,comStr):
        """
        爬取一页短评
        :param start:
        :return:
        """
        try:
            r = session.get(commentUrl,headers = comHeaders)
            r.raise_for_status()
        except:
            print('failure,start = %d' % start)

        doc = etree.HTML(r.text)
        self.total_author = doc.xpath('//div[@class="mod-bd"]/div[@class="comment-item"]/div[@class="comment"]/h3/span[@class="comment-info"]/a/text()')  # 作者
        self.total_star_num = doc.xpath('//div[@class="mod-bd"]/div[@class="comment-item"]/div[@class="comment"]/h3//span[@class="votes"]/text()')  # 有用数
        self.total_comment_time = doc.xpath('//div[@class="mod-bd"]/div[@class="comment-item"]/div[@class="comment"]/h3/span[@class="comment-info"]/span[3]/@title')  # 评论时间
        self.total_recommendation_level = doc.xpath('//div[@class="mod-bd"]/div[@class="comment-item"]/div[@class="comment"]/h3/span[@class="comment-info"]/span[2]/@title')  # 推荐程度
        self.total_text = doc.xpath('//div[@class="mod-bd"]//p//span/text()')  # 影评

        #写入excel
        if (self.write_excel(start) == 1):
            print("page %d finish" % start)
        else:
            print("page %d fail" % start)

    def write_excel(self,start):
        path = "../comments/"+str(start/20) +"_" + comStr + "_" + self.COMMENT_FILE_PATH
        try:
            column_headers = ["编号", "作者", "推荐程度", "评论时间", "点赞数", "详细影评"]
            if os.path.exists(path):
                os.remove(path)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.column_dimensions["A"].width = 5  # 列宽
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["F"].width = 700
            for i in range(6):  # 写列标题:["编号", "作者", "推荐程度", "评论时间", "点赞数", "详细影评"]
                ws.cell(row=1, column=i + 1, value=column_headers[i])
            for i in range(len(self.total_author)):
                ws.cell(row=i + 2, column=1, value=i + 1)  # 写编号这一列数据
                ws.cell(row=i + 2, column=2, value=self.total_author[i].encode("utf-8"))  # 作者
                ws.cell(row=i + 2, column=3, value=self.total_recommendation_level[i].encode("utf-8"))  # 推荐程度
                if i >= len(self.total_comment_time):pass
                else:ws.cell(row=i + 2, column=4, value=self.total_comment_time[i].encode("utf-8"))
                ws.cell(row=i + 2, column=5, value=self.total_star_num[i].encode("utf-8"))
                ws.cell(row=i + 2, column=6, value=self.total_text[i].encode("utf-8"))
            wb.save(path)
            print(path+" finish")
            return 1
        except Exception:
            print(traceback.print_exc())
            return 0

    def batch_get_comment(self,comStr):
        """
        批量爬取短评
        :return:
        """
        # comment pages properties
        comHeaders = {
            'User-Agent': 'Mozilla/5.0',
            'Referer': 'https://movie.douban.com/subject/25958787/',
            'Host': 'movie.douban.com'
        }
        eng = {
            '好评':'h',
            '一般':'m',
            '差评':'l'
        }
        for start in range(0,500,20):
            commentUrl = 'https://movie.douban.com/subject/25958787/comments?start={}&limit=20&sort=new_score&status=P&percent_type={}'.format(start,eng.get(comStr))
            self.getComment(commentUrl, comHeaders, start,comStr)
            time.sleep(random.random()*3)

if __name__ == '__main__':
    # login properties
    Html_headers = {'user-agent': 'Mozilla/5.0',
                    'Referer': 'https://movie.douban.com/subject/21937452/',
                    'Host': 'www.douban.com'
                    }
    cookies = {'cookie': 'bid=LBG6zC7_u4k; __yadk_uid=6o7lh2HbfchBzFuBeUO1PHaVThrMlxT9; douban-fav-remind=1; __gads=ID=f808c9c65786b98c:T=1563645029:S=ALNI_Mau-GHqEKAgk7G3pUt_E98z8Li98g; ll="118318"; _vwo_uuid_v2=D94AC1DDE037D44A987B719409539CCAA|0c47142066627434469eefc821521e0d; __utmc=30149280; __utmz=30149280.1571217856.6.6.utmcsr=google|utmccn=(organic)|utmcmd=organic|utmctr=(not%20provided); dbcl2="148384495:RAN55bcNNqo"; ck=A-CN; push_noty_num=0; push_doumail_num=0; __utmv=30149280.14838; __utma=30149280.80479015.1560330548.1571283155.1571298574.12; ap_v=0,6.0; _pk_ref.100001.8cb4=%5B%22%22%2C%22%22%2C1571300359%2C%22https%3A%2F%2Fmovie.douban.com%2Fsubject%2F21937452%2Fcomments%3Fstart%3D60%26limit%3D20%26sort%3Dnew_score%26status%3DP%22%5D; _pk_ses.100001.8cb4=*; __utmt=1; _pk_id.100001.8cb4=f1d3af8bd725a444.1560330505.8.1571300363.1571280607.; __utmb=30149280.4.10.1571298574'}
    login_url = 'https://www.douban.com/'
    D_Spider = DB_Film_review_Spider()
    session = requests.session()
    if(D_Spider.loginDou(login_url,cookies,Html_headers)):
        comStr = input("请输入评论类型(好评、一般、差评):")
        if(comStr != '好评' or comStr != '一般'or comStr != '差评'):
            comStr = input("乱tm输，请重新输入:")
        else:
            D_Spider.batch_get_comment(comStr)




